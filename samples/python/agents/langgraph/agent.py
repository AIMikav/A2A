import datetime
import os.path
import calendar
from typing import Any, Dict, AsyncIterable, Literal
from pydantic import BaseModel
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import dateparser
from openpyxl import Workbook, load_workbook
import logging

SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]

class ResponseFormat(BaseModel):
    status: Literal["input_required", "completed", "error"] = "input_required"
    message: str

class CalendarAgent:
    SYSTEM_INSTRUCTION = (
        "You are a specialized assistant for calendar tracking and day planning. "
        "Your purpose is to help users view and manage their Google Calendar events, "
        "summarize their day, and answer questions about their schedule. "
        "If the user asks about anything unrelated to calendar or planning, "
        "politely state that you can only assist with calendar-related queries. "
        "Set response status to input_required if the user needs to provide more information. "
        "Set response status to error if there is an error while processing the request. "
        "Set response status to completed if the request is complete."
    )

    SUPPORTED_CONTENT_TYPES = ["text", "text/plain"]

    def __init__(self):
        pass

    def _get_credentials(self):
        creds = None
        if os.path.exists("token.json"):
            creds = Credentials.from_authorized_user_file("token.json", SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    "credentials.json", SCOPES
                )
                creds = flow.run_local_server(port=0)
            with open("token.json", "w") as token:
                token.write(creds.to_json())
        return creds

    def _fetch_events(self, date: datetime.datetime = None, end_date: datetime.datetime = None, maxResults: int = 10):
        try:
            creds = self._get_credentials()
            service = build("calendar", "v3", credentials=creds)
            params = {
                "calendarId": "primary",
                "maxResults": maxResults,
                "singleEvents": True,
                "orderBy": "startTime",
            }
            if date and end_date:
                params["timeMin"] = date.isoformat()
                params["timeMax"] = end_date.isoformat()
            elif date:
                params["timeMin"] = date.isoformat()
            else:
                params["timeMin"] = datetime.datetime.now(tz=datetime.timezone.utc).isoformat()
            events_result = service.events().list(**params).execute()
            events = events_result.get("items", [])
            return events
        except HttpError as error:
            return {"error": f"An error occurred: {error}"}

    def _get_date_range_from_query(self, query: str):
        query_lower = query.lower()
        now = datetime.datetime.now(tz=datetime.timezone.utc)
        # Previous week
        if "last week" in query_lower or "previous week" in query_lower:
            # ISO week: Monday is 0
            last_monday = (now - datetime.timedelta(days=now.weekday() + 7)).replace(hour=0, minute=0, second=0, microsecond=0)
            last_sunday = last_monday + datetime.timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
            return last_monday, last_sunday
        # This week
        if "this week" in query_lower or "current week" in query_lower:
            monday = (now - datetime.timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            sunday = monday + datetime.timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
            return monday, sunday
        # Last month
        if "last month" in query_lower or "previous month" in query_lower:
            first_of_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_month = first_of_this_month - datetime.timedelta(days=1)
            first_of_last_month = last_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_of_last_month = last_month.replace(day=calendar.monthrange(last_month.year, last_month.month)[1], hour=23, minute=59, second=59, microsecond=999999)
            return first_of_last_month, last_of_last_month
        # This month
        if "this month" in query_lower or "current month" in query_lower:
            first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_of_month = now.replace(day=calendar.monthrange(now.year, now.month)[1], hour=23, minute=59, second=59, microsecond=999999)
            return first_of_month, last_of_month
        return None, None

    def _append_events_to_excel(self, events, filename="calendar_activities.xlsx"):
        headers = ["Date", "Summary", "Attachments", "Conference Links"]
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # Always use the first worksheet
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
        rows_before = ws.max_row
        for event in events:
            date = event["start"].get("dateTime", event["start"].get("date"))
            summary = event.get("summary", "(No Title)")
            attachments = ", ".join([
                att.get("title", "") + " " + att.get("fileUrl", "") for att in event.get("attachments", [])
            ])
            conference = event.get("conferenceData", {})
            entry_points = conference.get("entryPoints", [])
            conference_links = ", ".join([
                ep.get("uri", "") for ep in entry_points if ep.get("uri")
            ])
            ws.append([date, summary, attachments, conference_links])
        wb.save(filename)
        logging.info(f"Appended {ws.max_row - rows_before} rows to {filename}")

    def _plan_my_day(self, events):
        if not events:
            return "You have no events scheduled for today. Your day is wide open!"
        # Sort events by start time
        def parse_dt(dt):
            return dateparser.parse(dt)
        events_sorted = sorted(events, key=lambda e: parse_dt(e["start"].get("dateTime", e["start"].get("date"))))
        summary_lines = []
        summary_lines.append(f"You have {len(events_sorted)} event(s) today.")
        first_event = events_sorted[0]
        last_event = events_sorted[-1]
        first_time = first_event["start"].get("dateTime", first_event["start"].get("date"))
        last_time = last_event["end"].get("dateTime", last_event["end"].get("date"))
        summary_lines.append(f"First event: {first_time} - {first_event.get('summary', '(No Title)')}")
        summary_lines.append(f"Last event: {last_time} - {last_event.get('summary', '(No Title)')}")
        # Find free slots between events
        free_slots = []
        for i in range(len(events_sorted) - 1):
            end_current = parse_dt(events_sorted[i]["end"].get("dateTime", events_sorted[i]["end"].get("date")))
            start_next = parse_dt(events_sorted[i+1]["start"].get("dateTime", events_sorted[i+1]["start"].get("date")))
            if end_current < start_next:
                free_slots.append(f"Free from {end_current.strftime('%H:%M')} to {start_next.strftime('%H:%M')}")
        if free_slots:
            summary_lines.append("Suggested free slots for breaks or tasks:")
            summary_lines.extend(free_slots)
        else:
            summary_lines.append("No free slots between events. Consider scheduling breaks before your first or after your last event.")
        return "\n".join(summary_lines)

    def invoke(self, query, sessionId) -> Dict[str, Any]:
        try:
            # Check for 'plan my day' or similar queries
            if any(word in query.lower() for word in ["plan my day", "organize my day", "what's my schedule", "day at a glance"]):
                today = datetime.datetime.now(tz=datetime.timezone.utc)
                date = today.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
                events = self._fetch_events(date, end_date, maxResults=50)
                if isinstance(events, dict) and "error" in events:
                    return {
                        "is_task_complete": False,
                        "require_user_input": True,
                        "content": events["error"],
                    }
                self._append_events_to_excel(events)
                plan = self._plan_my_day(events)
                return {
                    "is_task_complete": True,
                    "require_user_input": False,
                    "content": plan,
                }
            # Check for 'all events' or 'history' queries
            if any(word in query.lower() for word in ["all events", "everything", "history", "full list"]):
                date = datetime.datetime(1970, 1, 1, tzinfo=datetime.timezone.utc)
                end_date = datetime.datetime.now(tz=datetime.timezone.utc)
                maxResults = 100
                events = self._fetch_events(date, end_date, maxResults=maxResults)
            else:
                # Try to parse a date range from the query
                date, end_date = self._get_date_range_from_query(query)
                maxResults = 10
                if not date and not end_date:
                    # Try to parse a single date from the query
                    parsed_date = dateparser.parse(query, settings={"TIMEZONE": "UTC", "RETURN_AS_TIMEZONE_AWARE": True})
                    if parsed_date:
                        date = parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)
                        end_date = parsed_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                    elif "tomorrow" in query.lower():
                        date = datetime.datetime.now(tz=datetime.timezone.utc) + datetime.timedelta(days=1)
                        date = date.replace(hour=0, minute=0, second=0, microsecond=0)
                        end_date = date.replace(hour=23, minute=59, second=59, microsecond=999999)
                    elif "today" in query.lower() or "next" in query.lower():
                        date = datetime.datetime.now(tz=datetime.timezone.utc)
                        date = date.replace(hour=0, minute=0, second=0, microsecond=0)
                        end_date = date.replace(hour=23, minute=59, second=59, microsecond=999999)
                events = self._fetch_events(date, end_date, maxResults=maxResults)
            if isinstance(events, dict) and "error" in events:
                return {
                    "is_task_complete": False,
                    "require_user_input": True,
                    "content": events["error"],
                }
            if not events:
                if 'maxResults' in locals() and maxResults == 100:
                    return {
                        "is_task_complete": True,
                        "require_user_input": False,
                        "content": "No events found in your calendar history.",
                    }
                if date and end_date:
                    if date == end_date:
                        return {
                            "is_task_complete": True,
                            "require_user_input": False,
                            "content": f"No events found for {date.strftime('%Y-%m-%d')}",
                        }
                    else:
                        return {
                            "is_task_complete": True,
                            "require_user_input": False,
                            "content": f"No events found from {date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                        }
                return {
                    "is_task_complete": True,
                    "require_user_input": False,
                    "content": "No upcoming events found.",
                }
            self._append_events_to_excel(events)
            summary = self._summarize_events(events)
            return {
                "is_task_complete": True,
                "require_user_input": False,
                "content": summary,
            }
        except Exception as e:
            return {
                "is_task_complete": False,
                "require_user_input": True,
                "content": f"An error occurred: {e}",
            }

    async def stream(self, query, sessionId) -> AsyncIterable[Dict[str, Any]]:
        yield {"is_task_complete": False, "require_user_input": False, "content": "Looking up your calendar events..."}
        result = self.invoke(query, sessionId)
        yield result

    def _summarize_events(self, events):
        lines = []
        for event in events:
            start = event["start"].get("dateTime", event["start"].get("date"))
            summary = event.get("summary", "(No Title)")
            attachments = event.get("attachments", [])
            attachment_lines = []
            for att in attachments:
                title = att.get("title", "(Attachment)")
                url = att.get("fileUrl", "")
                attachment_lines.append(f"Attachment: {title} {url}")
            # Conference/recording links
            conference = event.get("conferenceData", {})
            entry_points = conference.get("entryPoints", [])
            conference_lines = []
            for ep in entry_points:
                if ep.get("entryPointType") in ["video", "phone", "more"]:
                    label = ep.get("label", ep.get("entryPointType"))
                    uri = ep.get("uri", "")
                    conference_lines.append(f"Conference: {label} {uri}")
            lines.append(f"{start}: {summary}")
            if attachment_lines:
                lines.extend(attachment_lines)
            if conference_lines:
                lines.extend(conference_lines)
        return "Here are your events:\n" + "\n".join(lines)
